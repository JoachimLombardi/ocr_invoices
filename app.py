import json
from pathlib import Path
from openai import OpenAI
import os
import fitz
import base64
import pandas as pd
from PIL import Image
import streamlit as st
import tempfile
from openpyxl import load_workbook, Workbook
from dateutil import parser
from dotenv import load_dotenv

load_dotenv()


def to_french_date(date_str: str) -> str:
    """
    Convert a date string to a French date string.

    Args:
        date_str (str): The date string to convert.

    Returns:
        str: The French date string.
    """
    try:
        dt = parser.parse(date_str, dayfirst=False) 
        return dt.strftime("%d/%m/%Y")
    except Exception as e:
        print(f"Impossible de parser la date {date_str}: {e}")
        return date_str 


def invoice_to_image_and_text(invoice):
    """
    Convert a PDF document to an image and a text.

    Args:
        path (str): The path to the PDF document.
        
    Returns:
        dict or Exception: The text and image, or an Exception if the API call fails.
    """
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
        tmp_file.write(invoice.read())
        path = tmp_file.name
    path = Path(path)
    doc = fitz.open(path)
    list_img_path = []
    text = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        # Extract text
        text.append(page.get_text())
        # Image path
        img_path = path.parent / f"{path.stem}_{page_num}.jpg"
        list_img_path.append(img_path)
        # Convert to image
        matrix = fitz.Matrix(2, 2) 
        pix = page.get_pixmap(matrix=matrix)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        # Save image
        img.save(img_path, "JPEG", quality=100)
    doc.close()
    try:
        list_b64 = []
        for img_path in list_img_path:
            with open(img_path, "rb") as image_file:
               list_b64.append(base64.b64encode(image_file.read()).decode("utf-8"))
        return list_b64, text
    except Exception as e:
        print(f"The file {path} could not be read. Error: {e}")
        return None
    

def fill_excel_file(list_invoices_dict, csv_file):
    """
    Fill an Excel file with the invoice data.
    
    Args:
        invoice_dict (dict): The invoice data.
        excel_path (str): The path to the Excel file.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file.write(csv_file.read())
        excel_path = tmp_file.name
    excel_path = Path(excel_path)
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()
    for invoice in list_invoices_dict:
        company_name = invoice.get("company_name", "Unknown")
        invoice_ref = invoice.get("invoice_reference", {})
        invoice_number = invoice_ref.get("invoice_number", "")
        invoice_date = invoice_ref.get("invoice_date_raw", "")
        invoice_date = to_french_date(invoice_date)
        invoice_num = invoice_number + "DU " + invoice_date
        if company_name in wb.sheetnames:
            ws = wb[company_name]
        else:
            ws = wb.create_sheet(title=company_name)
            # En-têtes
            headers = [
                "N° FACTURE", "REF",
                "Article", "Quantité facturée",
                "Prix unitaire", "Total payé HT"
            ]
            ws.append(headers)
        for article in invoice.get("articles", []):
            row = [
                invoice_num,
                article.get("reference", ""),
                article.get("designation", ""),
                article.get("quantity", ""),
                article.get("unit_price", ""),
                article.get("total_price", "")
            ]
            ws.append(row)
        wb.save(excel_path)


prompt = """
   You are an expert assistant.
    You are given:
    - the text extracted from a PDF invoice
    - an image of the invoice
    Your task:
    - extract the relevant invoice information
    - compare the extracted text with the image to validate correctness
    - call the function extract_invoice_data with the extracted values
    Do not hallucinate values.
    """

tools = [{
    "type":"function",
    "function":{
        "name": "extract_invoice_data",
        "description": "Extract structured data from an invoice",
        "parameters": {
            "type": "object",
            "properties": {
                "company_name": {
                    "type": "string",
                    "description": "Name of the company issuing the invoice"
                },
                "invoice_reference": {
                    "type": "object",
                    "description": "Invoice reference as written on the document",
                    "properties": {
                        "invoice_number": {
                            "type": "string",
                            "description": "Invoice identifier extracted from the invoice header"
                        },
                        "invoice_date_raw": {
                            "type": "string",
                            "description": "Invoice date exactly as written on the invoice (any format)"
                        },
                    },
                    "required": ["invoice_number", "invoice_date_raw"]
                },
                "articles": {
                    "type": "array",
                    "description": "List of items listed on the invoice",
                    "items": {
                        "type": "object",
                        "properties": {
                            "reference": {
                                "type": "string",
                                "description": "Item reference or SKU"
                            },
                            "designation": {
                                "type": "string",
                                "description": "Item name or description"
                            },
                            "quantity": {
                                "type": "number",
                                "description": "Quantity of the item"
                            },
                            "unit_price": {
                                "type": "number",
                                "description": "Unit price before tax"
                            },
                            "total_price": {
                                "type": "number",
                                "description": "Total price for this item (quantity × unit price)"
                            }
                        },
                        "required": ["reference", "designation", "quantity", "unit_price", "total_price"]
                    }
                },
            },
            "required": ["company_name", "invoice_reference", "articles"]
        }
    }
}]
st.title("Extraction factures")
invoices = st.file_uploader(
    "Factures",
    type=["pdf", "png", "jpg"],
    accept_multiple_files=True
)
csv_file = st.file_uploader("Fichier Excel", type=["csv", "xlsx"])
if st.button("Lancer le traitement"):
    if not invoices or not csv_file:
        st.error("Veuillez fournir des factures et un fichier Excel")
    else:
        list_invoices_dict = []
        for invoice in invoices:
            list_b64, list_texts = invoice_to_image_and_text(invoice)
            text = "\n".join(list_texts)
            messages = [{"role":"user", "content": prompt + "\n\n" + text}]
            data = {
                    "model": "openai/gpt-oss-20b",
                    "messages": messages,
                    "tools": tools,
                    "stream": False,
                    "temperature": 0,
                    "max_tokens": 5000,
                    "top_p": 1e-6,
                    "seed": 42
                    }
            for attempt in range(1,4):
                try:
                    print("api huggingface llm call")
                    client = OpenAI(base_url="https://router.huggingface.co/v1",
                                    api_key=os.getenv("HUGGINGFACE_API_KEY"))
                    response = client.chat.completions.create(**data) 
                    final_response = response.choices[0].message
                    tool_call = None
                    if hasattr(final_response, "tool_calls") and final_response.tool_calls:
                        tool_call = final_response.tool_calls[0]
                    elif isinstance(final_response, dict) and final_response.get("tool_calls"):
                        tool_call = final_response["tool_calls"][0]
                    invoice_dict = None
                    if tool_call:
                        if isinstance(tool_call, dict):
                            invoice_dict = tool_call.get("function", {}).get("arguments", None)
                        else:
                            invoice_dict = getattr(tool_call.function, "arguments", None)
                    if isinstance(invoice_dict, dict):
                        try:
                            invoice_dict = json.loads(invoice_dict)
                        except json.JSONDecodeError:
                            pass
                    list_invoices_dict.append(invoice_dict)
                    break
                except Exception as e:
                    print(f"Attempt {attempt}/3 \n API call failed with error: {e} - retrying...")
        fill_excel_file(list_invoices_dict, csv_file)






        
